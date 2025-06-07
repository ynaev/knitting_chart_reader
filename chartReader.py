from openpyxl import load_workbook
import os
import glob
import tkinter as tk

def main():
    '''
    Reads an excel chart in working routine and displays each cell in turn
    ** User taps the screen to advance to the next cell
    ** Displays how many stitches should currently be on each needle
    ** Displays the current row
    '''
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    popup = tk.Toplevel(root)
    popup.title("Current Stitch")
    popup.geometry(f"600x250+50-50")

    # Add and store label for later updates
    label = tk.Label(
        popup,
        text='',
        font=("Consolas", 14),
        padx=10,
        pady=10,
        justify="center",
        bg="#fff9c4",
        fg="#000000"
    )
    label.pack()
    popup.label = label  # So you can update text later with popup.label.config(...)
    popup.withdraw()

    user_pressed_key = tk.BooleanVar()
#==============================================================================================
    def workRight(n, rounds, rc):
        stLeftNeedle = 0 #verify all are stitches on the needle
        stRightNeedle = nos  #verify there are no stitches on the other needle
        l = 0  #position the column counter
            
        while l <= nos - 1:
            c = ws[n][l]  #calibrate stitch position, then calculate the new stitch count
            cell_value = c.value
            if stRightNeedle < 0:
                raise ValueError(f"Negative stitch count! Row {rc}, col {l}, stitch {cell_value}")
            elif cell_value == 'K':
                stLeftNeedle += 1
                stRightNeedle -= 1
                l += 1
            elif cell_value == 'YO':
                stLeftNeedle += 1
                l += 1
            elif cell_value in ('K2T', 'SSK'):
                stLeftNeedle += 1
                stRightNeedle -= 2
                l += 2
            elif cell_value == 'KFB':
                stLeftNeedle += 2
                stRightNeedle -= 1
                l += 1

            if stRightNeedle > 0:
                show_stitch_popup(cell_value, stLeftNeedle, stRightNeedle, rc, rounds)
            elif stRightNeedle == 0:
                row_finished_popup(cell_value, stLeftNeedle, stRightNeedle, rc, rounds)

#==============================================================================================
    def workLeft(n, rounds, rc):
        stRightNeedle = 0 #verify all stitches are on the needle
        stLeftNeedle = nos  #verify there are no stitches on the other needle
        l = nos - 1  #position the column counter

        while l >= 0:
            c = ws[n][l]  #verify stitch position
            cell_value = c.value
            if stLeftNeedle < 0:
                raise ValueError(f"Negative stitch count! Row {rc}, col {l}, stitch {cell_value}")
            elif cell_value == 'K':
                stRightNeedle += 1
                stLeftNeedle -= 1
                l -= 1
            elif cell_value == 'YO':
                stRightNeedle += 1
                l -= 1
            elif cell_value in ('K2T', 'SSK'):
                stRightNeedle += 1
                stLeftNeedle -= 2
                l -= 2
            elif cell_value == 'KFB':
                stRightNeedle += 2
                stLeftNeedle -= 1
                l -= 1

            if stLeftNeedle > 0:
                show_stitch_popup(cell_value, stLeftNeedle, stRightNeedle, rc, rounds)
            elif stLeftNeedle == 0:
                row_finished_popup(cell_value, stLeftNeedle, stRightNeedle, rc, rounds)
        
#==============================================================================================
    def even_first(rounds):
        n = rir
        rc = 1

        while n > 1:  #start a new row
            if(n % 2) == 0:
                workLeft(n, rounds, rc)
            elif(n % 2) != 0:
                workRight(n, rounds, rc)
            n -= 1
            rc += 1

        if n == 1: #start last row in current round
            workRight(n, rounds, rc)

#==============================================================================================
    def odd_first(rounds):
        n = rir
        rc = 1

        while n > 1:  #start a new row
            if(n % 2) != 0:
                workLeft(n, rounds, rc)
            elif(n % 2) == 0:
                workRight(n, rounds, rc)
            n -= 1
            rc += 1

        if n == 1: #start last row in current round
            workLeft(n, rounds, rc)

#==============================================================================================
    def show_popup(text):
        label.config(text=text)
        popup.deiconify()
        popup.lift()
        popup.focus_force()
        popup.grab_set()

        user_pressed_key.set(False)
        popup.bind("<Key>", lambda e: user_pressed_key.set(True))
        popup.wait_variable(user_pressed_key)
        popup.unbind("<Key>")

        popup.withdraw()

#==============================================================================================
    def show_stitch_popup(cell_value, stLeftNeedle, stRightNeedle, rc, rounds):
        text=(
            f"🧵 Stitch: {cell_value} 🧵\n\n"
            f"🧶 Left needle: {stLeftNeedle} stitches\n"
            f"🧶 Right needle: {stRightNeedle} stitches\n\n"
            f"🔁 Row {rc} of Round {rounds}\n\n"
            "Press any key to continue..."
        )        
        show_popup(text)

#==============================================================================================
    def row_finished_popup(cell_value, stLeftNeedle, stRightNeedle, rc, rounds):
        text=(
            f"🧵Stitch: {cell_value}🧵\n\n"
            f"🧶Left needle should have {stLeftNeedle} stitches.\n"
            f"🧶Right needle should have {stRightNeedle} stitches.\n\n"                
            f"➡️Row {rc} of round {rounds} complete!  Moving on to row {rc + 1}...\n\n"
            "Press any key to continue..."
        )
        show_popup(text)

#==============================================================================================
    def round_finished_popup(rounds):
        text=(
            f"🔁Round {rounds} complete! Preparing for round {rounds + 1}\n\n"
            "Press any key to continue..."
        )
        show_popup(text)

#==============================================================================================
    def final_round_finished_popup():
        text=(
            f"✨ Project complete! ✨\n\n"
            "Press any key to continue..."
        )
        show_popup(text)

#==============================================================================================
    def choose_excel_file():
        # Get all .xlsx files in current folder
        folder = os.path.dirname(os.path.abspath(__file__))
        excel_files = glob.glob(os.path.join(folder, "*.xlsx"))

        if not excel_files:
            print("😢 No Excel files found in this folder.")
            exit()

        print("📄 Excel files found:")
        for idx, file in enumerate(excel_files, start=1):
            print(f"{idx}. {os.path.basename(file)}")

        while True:
            try:
                choice = int(input("\n🔢 Choose a file by number: "))
                if 1 <= choice <= len(excel_files):
                    return excel_files[choice - 1]
                else:
                    print("⚠️ Number out of range. Try again.")
            except ValueError:
                print("❌ Not a number. Try again.")

#==============================================================================================
    filename = choose_excel_file()
    wb = load_workbook(filename=filename)
    ws = wb.active

    nos = int(ws.max_column)  #number of stitches
    rir = int(ws.max_row)  #number of rows in a round

#==============================================================================================
    while True:
        try:
            nor = int(input("🧶 How many rounds would you like to complete? "))
            if nor < 1:
                print("Let's try a number greater than 0, yeah?")
                continue
            break
        except ValueError:
            print("That doesn't look like a number. Try again, preferably with digits this time.")
    
    even_rows = (rir % 2 == 0) #determine whether there are an even number of rows
    rounds = 1 #set the round counter to begin the project

    if even_rows:
        while rounds < nor: #begin the project/start a new round
            even_first(rounds)
            round_finished_popup(rounds)
            rounds += 1
        
        if rounds == nor: #begin the last round of the project
            even_first(rounds)
            final_round_finished_popup()
    
    elif not even_rows:
        while rounds < nor: #begin the project/start a new round
            odd_first(rounds)
            round_finished_popup(rounds)
            rounds += 1
        
        if rounds == nor: #begin the last round of the project
            odd_first(rounds)
            final_round_finished_popup()

if __name__ == "__main__":
    main()